
import tkinter  as tk 
from tkinter import * 
from subprocess import call
import pandas as pd
import openpyxl

my_w = tk.Tk()
#my_w.geometry("800x600") 
my_w.title("VISITOR'S MANAGEMENT SYSTEM: ")

path = "D:\\Visitors_managenemt_sys-master\\Students.xlsx"

wb = openpyxl.load_workbook(path)

sheet_obj = wb.active

max_col = sheet_obj.max_column
max_row = sheet_obj.max_row


e=Label(my_w,width=20,text='Id',borderwidth=10, relief='ridge',anchor=CENTER,bg='yellow')
e.grid(row=1,column=1)
e=Label(my_w,width=20,text='Name',borderwidth=10, relief='ridge',anchor=CENTER,bg='yellow')
e.grid(row=1,column=2)
e=Label(my_w,width=20,text='phone NO',borderwidth=10, relief='ridge',anchor=CENTER,bg='yellow')
e.grid(row=1,column=3)
e=Label(my_w,width=20,text='Allow',borderwidth=10, relief='ridge',anchor=CENTER,bg='yellow')
e.grid(row=1,column=4)
e=Label(my_w,width=20,text='Wait',borderwidth=10, relief='ridge',anchor=CENTER,bg='yellow')
e.grid(row=1,column=5)

variable = 1
def make_something(value):
    global variable
    variable = value
    
def open_py_file():
    #call(["python","classifier.py"])
    call(["python","dataGathering.py"])

b3=Button(fg="white",text="Available",width=20,bg="green")
b3.grid(row=3,column=6)
b4=Button(fg="white",text="Not Available",width=20,bg="red")
b4.grid(row=5,column=6)

i=2
j=1
for i in range(2,max_row+1): 
    for j in range(1,max_col+1):
        cell_obj = sheet_obj.cell(row = i, column = j)
        e = Label(my_w,width=20, text=cell_obj.value,borderwidth=2,relief='ridge', anchor=CENTER,bg="orange")  
        e.grid(row=i, column=j) 
        #e.insert(END, visitor[j])
        
    
    b1 = Button(fg="red",text="Allow",width=10,bg="light blue",command=lambda *args: make_something(j))#need to add Command= to execute any function(backend part)
    b1.grid(row=i,column=j+1)
    b2 = Button(fg="red",text="Wait",width=10,bg="light blue",command=lambda *args: make_something(j))#need to add Command= to execute any function(backend part)
    b2.grid(row=i,column=j+2)
    i=i+1


for i in range(1, max_row+1):
    for j in range(1,max_col+1):
        cell_obj = sheet_obj.cell(row = i, column = j)
        print(cell_obj.value, end = "   ")
      

    print("\n")

 
my_w.mainloop()